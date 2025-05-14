import os
from typing import Type, Any, Dict
from ExcelTools import 表格工具
from Style.StyleDefiner import 表格样式类型
from TableData import 关卡表, 怪物基础数据表, 肉鸽技能节奏, 角色成长表, 阵容战力成长时间分布, 阵容战力成长表
from LubanData import 全局参数
# from MonsterData.MonsterDataManager import 怪物数据保存器
from MonsterData.MonsterTypes import 怪物设计表
from MonsterData.MonsterDataGenerater import 怪物设计理论值基本参数, 计算怪物属性
# 终端清屏，快速查看bug
os.system('cls' if os.name == 'nt' else 'clear')

# -------------------------------------
# 1. 定义全局变量
# -------------------------------------
# 保存文件夹的地址
save_folder_path = r"H:\悠手好闲\铁头工作室\方块项目设计案\《冲王》数值_cursor\新数据生成"
save_folder_path_level = r"H:\悠手好闲\铁头工作室\方块项目设计案\《冲王》数值_cursor\新数据生成\关卡表"

# 特殊行和特殊列的定义，这个项目是luban的导入的格式需要的
特殊行数 = 3  # 前3行是特殊行（##var, ##type, ##）
特殊列数 = 1  # 第1列是特殊列

# -------------------------------------
# 2. 生成表格的定义函数
# -------------------------------------
class 生成表:
    @classmethod
    def 从列函数映射提取表头(cls, 列函数映射: Dict[int, Any]) -> Dict[str, str]:
        """从列函数映射中提取表头信息
        
        Args:
            列函数映射: 列函数映射字典
            
        Returns:
            Dict[str, str]: 表头字典，键为列名，值为数据类型
        """
        表头字典 = {}
        for _, 列定义 in 列函数映射.items():
            # 列名作为表头字典的键
            列名 = 列定义.列名
            
            # 检查列名是否包含特殊符号，这部分逻辑来自表头信息类
            特殊符号 = "()（）/\\|<>{}[]【】《》"
            for 符号 in 特殊符号:
                if 符号 in 列名:
                    raise ValueError(f"表头显示名称 '{列名}' 不能包含特殊符号 '{符号}'")
            
            # 使用列定义中的表头数据类型
            表头字典[列名] = 列定义.表头数据类型
        return 表头字典

    @classmethod
    def 生成表格(cls, 表格定义: Type[Any], sheet_name: str = None, save_folder: str = None, 
                表格样式: 表格样式类型 = 表格样式类型.默认样式) -> None:
        """生成指定类型的表格
        根据表格定义类生成对应的Excel表格
        Args:
            表格定义: 表格定义类，包含表格的所有数据定义
            sheet_name: 工作表名称，如果为None则使用表格定义中的默认名称
            save_folder: 保存Excel文件的文件夹路径，如果为None则使用默认路径
            表格样式: 表格的样式类型
        """
        # 使用默认值
        if save_folder is None:
            save_folder = save_folder_path
        if sheet_name is None:
            sheet_name = 表格定义.默认工作表名称

        # 创建保存目录
        os.makedirs(save_folder, exist_ok=True)
        file_path = os.path.join(save_folder, 表格定义.表格名称)

        # 获取工作表对象和样式
        wb, ws, 表头样式名, 数据样式名 = 表格工具.更新或创建工作表(
            file_path, sheet_name, 表格样式
        )

        # 写入特殊的第一列数据
        for i in range(特殊行数):
            ws.cell(row=i+1, column=1, value=f"##{'var' if i==0 else 'type' if i==1 else ''}").style = 表头样式名

        # 获取列函数映射
        column_functions = 表格定义.获取列函数映射()
        
        # 从列函数映射中提取表头，而不是使用预定义的表头枚举
        headers = cls.从列函数映射提取表头(column_functions)
        
        # 生成表头，从特殊列之后开始
        表格工具.生成表头(ws, headers, 表头样式名, start_row=1, start_col=特殊列数+1)

        # 获取行数据范围
        起始值, 结束值 = 表格定义.获取行数据范围()

        # 生成数据行，从特殊行之后开始
        for 当前值 in range(起始值, 结束值 + 1):
            row_num = 当前值 - 起始值 + 特殊行数 + 1  # +1是因为行号从1开始
            
            # 创建行参数，等于创建了实际键的映射表
            row_params = 表格定义.创建行参数(ws, ws.title, 当前值, row_num)

            # 创建列名到列号的映射
            列号映射 = {列名: 列号 for 列号, 列名 in enumerate(headers, 特殊列数+1)}  # 从特殊列之后开始

            # 生成每列数据
            for col_num, 列定义 in column_functions.items():
                # 根据参数映射创建实际的参数字典，等于重新定义了字典
                实际参数 = {映射键: row_params[实际键] for 映射键, 实际键 in 列定义.参数映射.items()}
                
                # 计算并写入单元格值
                raw_value = 列定义.计算值(实际参数, ws, row_num, 列号映射, 特殊行数+1)  # 特殊行数+1是数据起始行号
                
                # 确保值是基本类型（字符串、数字等），而不是单元格对象
                if isinstance(raw_value, (int, float)):
                    # 根据列名决定保留的小数位数
                    if "倍率" in 列定义.列名:
                        cell_value = round(float(raw_value), 4)
                    else:
                        cell_value = round(float(raw_value), 1)
                else:
                    cell_value = str(raw_value)
                
                cell = ws.cell(row=row_num, column=col_num + 特殊列数, value=cell_value)  # +特殊列数是因为前面有特殊列
                cell.style = 数据样式名

        # 调整列宽并保存文件
        表格工具.按中文调整列宽(ws, 特殊行数, 特殊列数+1, len(headers) + 特殊列数)  # 从特殊列之后开始
        wb.save(file_path)
        print(f"Excel文件已生成在：{file_path}")

    @classmethod
    def 将公式转换为值(cls, file_path: str) -> None:
        """将Excel文件中的所有公式转换为实际值"""
        from openpyxl import load_workbook
        # 加载工作簿，使用data_only=True来获取公式计算后的值
        wb = load_workbook(file_path, data_only=True)
        # 保存文件，这会覆盖原文件，但只包含值而不包含公式
        wb.save(file_path)
        print(f"已将公式转换为值：{file_path}")

    @classmethod
    def 生成所有等级的阵容战力成长时间分布表(cls, save_folder: str = save_folder_path, 
                                表格样式: 表格样式类型 = 表格样式类型.默认样式) -> None:
        """生成从1级到角色等级上限的所有等级的阵容战力成长时间分布表
        
        Args:
            save_folder: 保存Excel文件的文件夹路径，如果为None则使用默认路径
            表格样式: 表格的样式类型
        """
            
        # 从全局参数获取角色等级上限
        角色等级上限 = 全局参数.角色等级上限
        
        print(f"\n开始生成1级到{角色等级上限}级的阵容战力成长时间分布表...")
        
        # 循环生成每个等级的表格
        for 等级 in range(1, 角色等级上限 + 1):
            print(f"\n正在生成等级{等级}的阵容战力成长时间分布表...")
            cls.生成表格(
                表格定义=阵容战力成长时间分布,
                sheet_name=f"等级{等级}",
                save_folder=save_folder,
                表格样式=表格样式
            )
        
        print(f"\n所有等级(1-{角色等级上限})的阵容战力成长时间分布表生成完成!")

    @classmethod
    def 生成所有已设计怪物表格(cls, save_folder: str = save_folder_path, 
                            表格样式: 表格样式类型 = 表格样式类型.默认样式) -> None:
        """生成所有已设计怪物的表格，所有怪物数据放在同一个sheet中
        
        Args:
            save_folder: 保存Excel文件的文件夹路径，如果为None则使用默认路径
            表格样式: 表格的样式类型
        """
        # 初始化保存路径和文件名
        os.makedirs(save_folder, exist_ok=True)
        file_path = os.path.join(save_folder, 怪物基础数据表.表格名称)
        sheet_name = "所有怪物数据"
        
        print(f"\n开始生成所有已设计怪物的表格...")
        
        # 获取工作表对象和样式
        wb, ws, 表头样式名, 数据样式名 = 表格工具.更新或创建工作表(
            file_path, sheet_name, 表格样式
        )
        
        # 写入特殊的第一列数据
        for i in range(特殊行数):
            ws.cell(row=i+1, column=1, value=f"##{'var' if i==0 else 'type' if i==1 else ''}").style = 表头样式名
        
        # 获取列函数映射
        column_functions = 怪物基础数据表.获取列函数映射()
        
        # 从列函数映射中提取表头
        headers = cls.从列函数映射提取表头(column_functions)
        
        # 生成表头，从特殊列之后开始
        表格工具.生成表头(ws, headers, 表头样式名, start_row=1, start_col=特殊列数+1)
        
        # 获取行数据范围
        起始值, 结束值 = 怪物基础数据表.获取行数据范围()
        
        # 获取所有怪物数据
        print("正在获取所有怪物数据...")
        所有怪物数据 = 计算怪物属性.生成所有怪物实际数据()
        print(f"获取到 {len(所有怪物数据)} 个怪物数据")
        
        # 遍历所有怪物
        当前行号 = 特殊行数 + 1  # 从特殊行之后开始
        
        for 怪物 in 所有怪物数据:
            print(f"\n正在添加{怪物.名称}的数据...")
            # print(f"怪物属性: {怪物.__dict__}")
            
            # 遍历每个怪物的所有等级
            for 等级 in range(起始值, 结束值 + 1):
                try:
                    # 创建行参数
                    row_params = 怪物基础数据表.创建行参数(ws, 怪物.名称, 等级, 当前行号)
                    
                    # 创建列名到列号的映射
                    列号映射 = {列名: 列号 for 列号, 列名 in enumerate(headers, 特殊列数+1)}  # 从特殊列之后开始
                    
                    # 生成每列数据
                    for col_num, 列定义 in column_functions.items():
                        # 根据参数映射创建实际的参数字典
                        实际参数 = {映射键: row_params[实际键] for 映射键, 实际键 in 列定义.参数映射.items()}
                        
                        # 计算并写入单元格值
                        raw_value = 列定义.计算值(实际参数, ws, 当前行号, 列号映射, 特殊行数+1)
                        
                        # 确保值是基本类型（字符串、数字等），而不是单元格对象
                        if isinstance(raw_value, (int, float)):
                            # 根据列名决定保留的小数位数
                            if "倍率" in 列定义.列名:
                                cell_value = round(float(raw_value), 4)
                            else:
                                cell_value = round(float(raw_value), 1)
                        else:
                            cell_value = str(raw_value)
                        
                        cell = ws.cell(row=当前行号, column=col_num + 特殊列数, value=cell_value)
                        cell.style = 数据样式名
                    
                    当前行号 += 1
                except Exception as e:
                    print(f"处理怪物 {怪物.名称} 等级 {等级} 时出错: {str(e)}")
                    raise
        
        # 调整列宽并保存文件
        表格工具.按中文调整列宽(ws, 特殊行数, 特殊列数+1, len(headers) + 特殊列数)
        wb.save(file_path)
        print(f"\n所有已设计怪物的表格生成完成，文件已保存至：{file_path}")

    @classmethod
    def 生成关卡表(cls, 表格定义: Type[Any], 关卡编号: int, 关卡数据=None, save_folder: str = save_folder_path_level, 表格样式: 表格样式类型 = 表格样式类型.默认样式) -> None:
        """生成关卡设计表
        
        Args:
            表格定义: 表格定义类，包含表格的所有数据定义
            关卡编号: 关卡编号
            关卡数据: 关卡数据对象，如果提供则直接使用，否则需要重新生成
            save_folder: 保存Excel文件的文件夹路径
            表格样式: 表格的样式类型
        """
        # 初始化保存路径和文件名
        os.makedirs(save_folder, exist_ok=True)
        file_path = os.path.join(save_folder, f"关卡编号{关卡编号}.xlsx")
        sheet_name = f"{关卡编号}"
        
        print(f"\n开始生成关卡编号{关卡编号}的表格...")
        
        # 获取工作表对象和样式
        wb, ws, 表头样式名, 数据样式名 = 表格工具.更新或创建工作表(
            file_path, sheet_name, 表格样式
        )

        特殊行数 = 6
        特殊列数 = 3
        # 写入特殊数据
        # 定义单元格值的字典
        单元格值 = {
            (1, 1): "##var", (2, 1): "##type", (3, 1): "##var", (4, 1): "##", (5, 1): "##var", (6, 1): "##",
            (1, 2): "level_id", (1, 3): "time_before_spawn", (1, 5): "*wave_list",
            (2, 2): "int", (2, 3): "float", (2, 5): "list,WaveData",
            (3, 5): "*enemy_group_list",
            (4, 2): "关卡编号", (4, 3): "生成延迟时间", (4, 4): "波次", (4, 5): "组列表",
            (5, 5): "enemy_id", (5, 6): "enemy_count", (5, 7): "current_level",
        }
        
        # 合并单元格
        合并区域 = [
            'E1:G1', 'E2:G2', 'E3:G3', 'E4:G4', 'H1:J5', 'B6:C6'
        ]
        表头行范围 = 7
        表头列范围 = 14


        # 统一设置样式和值
        for 行 in range(1, 表头行范围):  # 1到6行
            for 列 in range(1, 表头列范围):  # 1到14列(A到M)
                if (行, 列) in 单元格值:
                    ws.cell(row=行, column=列, value=单元格值[(行, 列)]).style = 表头样式名
                else:
                    # 对于没有特定值的单元格，仍然设置样式
                    ws.cell(row=行, column=列).style = 表头样式名
        
        # 合并单元格
        for 区域 in 合并区域:
            ws.merge_cells(区域)
        
        ws.cell(row=7, column=2, value=关卡编号).style = 数据样式名
        ws.cell(row=7, column=3, value=10).style = 数据样式名

        # 获取列函数映射
        column_functions = 表格定义.获取列函数映射()
        
        # 从列函数映射中提取表头
        headers = cls.从列函数映射提取表头(column_functions)

        表格工具.生成关卡表头(ws, headers, 表头样式名)
        
        # 获取行数据范围
        起始值, 结束值 = 表格定义.获取行数据范围()
        
        # 生成数据行，从特殊行之后开始
        当前行号 = 特殊行数 + 1  # 从第7行开始
        
        # 遍历行数据范围
        for 当前值 in range(起始值, 结束值 + 1):
            try:
                # 创建行参数，传入关卡数据
                row_params = 表格定义.创建行参数(ws, sheet_name, 当前值, 当前行号, 关卡数据)
                
                # 创建列名到列号的映射
                列号映射 = {列名: 列号 for 列号, 列名 in enumerate(headers, 特殊列数+1)}  # 从特殊列之后开始
                
                # 生成每列数据
                for col_num, 列定义 in column_functions.items():
                    # 根据参数映射创建实际的参数字典
                    实际参数 = {映射键: row_params[实际键] for 映射键, 实际键 in 列定义.参数映射.items()}
                    
                    # 计算并写入单元格值
                    raw_value = 列定义.计算值(实际参数, ws, 当前行号, 列号映射, 特殊行数+1)
                    
                    # 确保值是基本类型（字符串、数字等），而不是单元格对象
                    if isinstance(raw_value, (int, float)):
                        # 根据列名决定保留的小数位数
                        if raw_value is None or raw_value == 0:
                            cell_value = ""
                        elif "倍率" in 列定义.列名:
                            cell_value = round(float(raw_value), 4)
                        else:
                            cell_value = round(float(raw_value), 1)
                    else:
                        cell_value = str(raw_value) if raw_value is not None else ""
                    
                    cell = ws.cell(row=当前行号, column=col_num + 特殊列数, value=cell_value)
                    cell.style = 数据样式名
                
                当前行号 += 1
            except Exception as e:
                print(f"处理关卡 {关卡编号} 行 {当前值} 时出错: {str(e)}")
                raise
        
        # 调整列宽并保存文件
        表格工具.调整关卡表列宽(ws, 特殊行数, 1, 表头列范围)
        wb.save(file_path)
        print(f"Excel文件已生成在：{file_path}")

    @classmethod
    def 生成所有关卡表格(cls, save_folder: str = save_folder_path_level, 表格样式: 表格样式类型 = 表格样式类型.默认样式) -> None:
        """生成所有关卡设计表数据的表格，每个关卡创建一个以关卡编号为名称的xlsx文件
        
        Args:
            save_folder: 保存Excel文件的文件夹路径
            表格样式: 表格的样式类型
        """
        from LevelData.LevelDataGenerator import 关卡生成器
        from TableData import 关卡表
        import time
        
        # 记录开始时间
        开始时间 = time.time()
        
        # 只生成一次所有关卡数据
        print("正在生成所有关卡数据...")
        所有关卡数据 = 关卡生成器.生成关卡设计表数据()
        print(f"共获取到 {len(所有关卡数据)} 个关卡数据")
        
        # 创建关卡ID到关卡数据的映射，便于快速查找
        关卡数据映射 = {关卡.关卡ID: 关卡 for 关卡 in 所有关卡数据}
        
        # 记录数据生成时间
        数据生成时间 = time.time() - 开始时间
        print(f"生成所有关卡数据耗时: {数据生成时间:.2f}秒")
        
        # 遍历所有关卡数据，为每个关卡生成一个表格
        for 关卡编号, 关卡数据 in 关卡数据映射.items():
            表格生成开始时间 = time.time()
            print(f"正在生成关卡{关卡编号}的表格...")
            cls.生成关卡表(关卡表, 关卡编号, 关卡数据, save_folder, 表格样式)
            表格生成时间 = time.time() - 表格生成开始时间
            print(f"生成关卡{关卡编号}表格耗时: {表格生成时间:.2f}秒")
        
        # 记录总时间
        总时间 = time.time() - 开始时间
        print(f"所有关卡表格生成完成！总耗时: {总时间:.2f}秒")


# 调用示例
if __name__ == "__main__":
    print("开始生成表格...")

    # 生成所有关卡表格
    生成表.生成所有关卡表格()

    # 生成单个关卡表格示例
    # 生成表.生成关卡表(
    #     表格定义=关卡表,
    #     关卡编号=10010,
    #     save_folder=save_folder_path_level,
    #     表格样式=表格样式类型.默认样式
    # )
    # 生成表.生成所有已设计怪物表格()
    


    # print("\n正在生成阵容战力成长表...")
    # 生成表.生成表格(
    #     表格定义=阵容战力成长表,
    #     sheet_name="标准阵容列表",
    #     save_folder=save_folder_path,
    #     表格样式=表格样式类型.默认样式
    # )

    # 生成表.生成所有等级的阵容战力成长时间分布表()

    # 生成怪物初始属性汇总表
    # 生成表.生成怪物初始属性汇总表()
