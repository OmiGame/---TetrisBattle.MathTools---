import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from Style.StyleDefiner import 表格样式类型, 样式生成器
from typing import Any, Dict

class 表格工具:
    @staticmethod
    def 更新或创建工作表(file_path: str, sheet_name: str, 表格样式: 表格样式类型) -> tuple:
        """更新或创建工作表，并返回工作簿、工作表、表头样式名和数据样式名"""
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            # 如果工作表已存在，先删除它，并放在第一位
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            # 在索引0的位置创建新工作表
            ws = wb.create_sheet(sheet_name, 0)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

        # 注册表格样式
        表头样式名, 数据样式名 = 表格工具.注册表格样式(wb, 表格样式)

        return wb, ws, 表头样式名, 数据样式名

    @staticmethod
    def 生成表头(ws: Worksheet, headers: Dict[str, str], style_name: str, start_row: int = 1, start_col: int = 2) -> None:
        """生成表头
        生成三行表头：第一行是变量名，第二行是数据类型，第三行是显示名称
        
        Args:
            ws: 工作表对象
            headers: 表头字典，key为显示名称，value为数据类型
            style_name: 样式名称
            start_row: 起始行号
            start_col: 起始列号
        """
        # 生成第一行（变量名）
        for col, (header, _) in enumerate(headers.items(), start_col):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.style = style_name
            
        # 生成第二行（数据类型）
        for col, (_, data_type) in enumerate(headers.items(), start_col):
            cell = ws.cell(row=start_row + 1, column=col, value=data_type)
            cell.style = style_name
            
        # 生成第三行（显示名称）
        for col, (header, _) in enumerate(headers.items(), start_col):
            cell = ws.cell(row=start_row + 2, column=col, value=header)
            cell.style = style_name

    @staticmethod
    def 按中文调整列宽(ws, start_row , start_col, end_col):
        """处理数据的起始列：遍历所有行，找最大值"""
        first_column = get_column_letter(start_col)
        max_length = 0
        for cell in ws[first_column]:  # 遍历该列所有单元格
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length  # 保留最大值
        adjusted_width = (max_length + 2) * 2.2  # 调整系数为 2.2
        ws.column_dimensions[first_column].width = adjusted_width

        """处理其他列： 需要传入表头的起始行"""
        for col in range(start_col + 1, end_col + 1):
            cell = ws.cell(row= start_row, column=col)
            content_length = len(str(cell.value)) if cell.value else 0
            adjusted_width = (content_length + 2) * 2.2  # 同样调整系数
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def 列表元素字符串拼接(列表, 分隔符=","):
        """将列表元素依次写入指定单元格"""
        cell_value = 分隔符.join(str(item) for item in 列表)
        return cell_value

    @staticmethod
    def 找到类字段(目标类: type, 字段名称: str) -> Any:
        """根据字段名称获取类中的字段值
        Args:
            目标类: 要查找的类
            字段名称: 要查找的字段名称
        Returns:
            Any: 字段的值
        Raises:
            AttributeError: 如果字段不存在
        """
        if not hasattr(目标类, 字段名称):
            raise AttributeError(f"类 {目标类.__name__} 中不存在变量名 '{字段名称}'")
            
        return getattr(目标类, 字段名称)

    @classmethod
    def 注册表格样式(cls, wb, 表格样式: 表格样式类型 = 表格样式类型.默认样式):
        """根据表格样式注册对应样式"""
        # 获取样式名称
        header_style_name = 样式生成器.获取样式名称(表格样式, "header")
        data_style_name = 样式生成器.获取样式名称(表格样式, "data")

        # 强制清除旧样式
        样式生成器.清除样式缓存(wb, header_style_name)
        样式生成器.清除样式缓存(wb, data_style_name)

        # 获取新配置
        header_config = 样式生成器.获取样式配置(header_style_name)
        data_config = 样式生成器.获取样式配置(data_style_name)

        # 创建并注册新样式
        wb.add_named_style(样式生成器.创建样式(header_config))
        wb.add_named_style(样式生成器.创建样式(data_config))

        return header_style_name, data_style_name
    
    # @classmethod
    # def 找到类字段(cls, target_name: str):
    # # 获取类的所有变量（排除特殊属性）
    #     class_vars = vars(cls)
    #     valid_vars = {k: v for k, v in class_vars.items() if not k.startswith("__")}
    
    #     if target_name in valid_vars:
    #         return valid_vars[target_name]
    #     else:
    #         raise AttributeError(f"类 {cls.__name__} 中不存在变量名 '{target_name}'")

